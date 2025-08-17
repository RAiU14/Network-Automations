import os
import pandas as pd
import logging
from . import Cisco_IOS_XE
import re
from datetime import datetime, timedelta
from dateutil import parser
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import json  # for export_json

# ----- version banner so you can confirm the right module is loaded -----
PHASE2_VERSION = "Phase2-final-2025-08-17"
try:
    logging.info(f"[Data_to_Excel] Loaded module version: {PHASE2_VERSION}")
except Exception:
    pass
# -----------------------------------------------------------------------

def _unwrap_value(val):
    while isinstance(val, list) and len(val) == 1:
        val = val[0]
    return val

# This function is used to write/append values to excel.
def append_to_excel(ticket_number, data_list, file_path=None, column_order=None):
    logging.debug(f"Appending to Excel for ticket{ticket_number}")
    if column_order is None:
        column_order = [
            'File name', 'Host name', 'Model number', 'Serial number', 'Interface ip address',
            'Uptime', 'Current s/w version', 'Current SW EOS', 'Suggested s/w ver', 's/w release date',
            'Latest S/W version', 'Production s/w is deffered or not?', 'Last Reboot Reason', 'Any Debug?',
            'CPU Utilization', 'Total memory', 'Used memory', 'Free memory', 'Memory Utilization (%)',
            'Total flash memory', 'Used flash memory', 'Free flash memory', 'Used Flash (%)',
            'Fan status', 'Temperature status', 'PowerSupply status', 'Available Free Ports',
            'End-of-Sale Date: HW', 'Last Date of Support: HW', 'End of Routine Failure Analysis Date:  HW',
            'End of Vulnerability/Security Support: HW', 'End of SW Maintenance Releases Date: HW',
            'Any Half Duplex', 'Interface/Module Remark', 'Config Status', 'Config Save Date',
            'Critical logs', 'Remark'
        ]
    if file_path is None:
        # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # file_path = f"{ticket_number}_network_analysis_{timestamp}.xlsx"
        file_path = f"{ticket_number}_network_analysis.xlsx"

    formatted_data = []

    if isinstance(data_list, dict):
        data_list = [data_list]

    for data in data_list:
        if not isinstance(data, dict):
            logging.error(f"Skipping invalid data: {type(data)} for {ticket_number}.")
            continue
        if not data:
            continue
        data_length = 1
        for value in data.values():
            if isinstance(value, list) and len(value) > 0:
                data_length = len(value)
                break
        for i in range(data_length):
            row_data = {}
            for key in column_order:
                if key in data:
                    value = data[key]
                    # Unwrap the value here
                    if isinstance(value, list):
                        unwrapped_value = _unwrap_value(value[i] if i < len(value) else 'Not available')
                        row_data[key] = unwrapped_value
                    else:
                        unwrapped_value = _unwrap_value(value)
                        row_data[key] = unwrapped_value
                else:
                    row_data[key] = 'Not available'
            formatted_data.append(row_data)

    if not formatted_data:
        logging.warning("No data to write to Excel.")
        return None

    df = pd.DataFrame(formatted_data)
    df = df[column_order]

    # Ensure parent folder exists (controller creates it, but make this bulletproof)
    try:
        parent = os.path.dirname(file_path)
        if parent and not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
    except Exception as e:
        logging.warning(f"Could not ensure parent folder for Excel: {e}")

    # Write/append Excel
    try:
        if os.path.exists(file_path):
            logging.info(f"Appending to existing Excel file: {file_path}")
            existing_df = pd.read_excel(file_path)
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df.to_excel(file_path, index=False)
        else:
            logging.info(f"Creating new Excel file: {file_path}")
            df.to_excel(file_path, index=False)
        logging.debug(f"Successfully wrote {len(df)} rows to Excel file and saved in {file_path}")
    except Exception as e:
        logging.error(f"Error writing to Excel for case {ticket_number}: {str(e)}")
        return None

    # --- Option B: always style + enforce sheet order/names here ---
    try:
        process_and_style_excel(file_path)
    except Exception as e:
        logging.warning(f"Excel styling step failed inside append_to_excel: {e}")

    try:
        add_summary_sheet(file_path)  # Ensures 1) Preventive Maintanance, 2) Summary
    except Exception as e:
        logging.warning(f"Summary sheet step failed inside append_to_excel: {e}")

    return file_path

def export_json(ticket_number, data_list, file_path=None, column_order=None, coerce_percentages=True):
    """
    Export the same tabular data used for Excel to a JSON file (list of objects).
    This does not change any parsing logic; it mirrors 'append_to_excel' shaping.
    """
    logging.debug(f"Exporting JSON for ticket {ticket_number}")
    if column_order is None:
        column_order = [
            'File name', 'Host name', 'Model number', 'Serial number', 'Interface ip address',
            'Uptime', 'Current s/w version', 'Current SW EOS', 'Suggested s/w ver', 's/w release date',
            'Latest S/W version', 'Production s/w is deffered or not?', 'Last Reboot Reason', 'Any Debug?',
            'CPU Utilization', 'Total memory', 'Used memory', 'Free memory', 'Memory Utilization (%)',
            'Total flash memory', 'Used flash memory', 'Free flash memory', 'Used Flash (%)',
            'Fan status', 'Temperature status', 'PowerSupply status', 'Available Free Ports',
            'End-of-Sale Date: HW', 'Last Date of Support: HW', 'End of Routine Failure Analysis Date:  HW',
            'End of Vulnerability/Security Support: HW', 'End of SW Maintenance Releases Date: HW',
            'Any Half Duplex', 'Interface/Module Remark', 'Config Status', 'Config Save Date',
            'Critical logs', 'Remark'
        ]

    if file_path is None:
        file_path = f"{ticket_number}_network_analysis.json"

    # Normalize input shape (same as append_to_excel)
    formatted_rows = []
    if isinstance(data_list, dict):
        data_list = [data_list]

    for data in data_list:
        if not isinstance(data, dict) or not data:
            logging.warning(f"Skipping invalid/empty data item for JSON export: {type(data)}")
            continue

        # Determine row count (same rule as Excel)
        data_length = 1
        for v in data.values():
            if isinstance(v, list) and len(v) > 0:
                data_length = len(v)
                break

        for i in range(data_length):
            row_obj = {}
            for key in column_order:
                if key in data:
                    value = data[key]
                    # Keep behavior consistent with Excel shaping
                    if isinstance(value, list):
                        val = _unwrap_value(value[i] if i < len(value) else 'Not available')
                    else:
                        val = _unwrap_value(value)
                else:
                    val = 'Not available'
                row_obj[key] = val
            formatted_rows.append(row_obj)

    # Optional: coerce percentage-like strings to numeric fractions (mirrors Excel step)
    if coerce_percentages:
        percent_cols = ["CPU Utilization", "Memory Utilization (%)", "Used Flash (%)"]
        def _to_fraction(val):
            try:
                if isinstance(val, (int, float)):
                    return float(val) if val <= 1 else val  # keep if already fraction
                if isinstance(val, str):
                    s = val.strip()
                    m = re.match(r'^(\d+(?:\.\d+)?)\s*%$', s)
                    if m:
                        return float(m.group(1)) / 100.0
            except Exception:
                pass
            return val

        for row in formatted_rows:
            for c in percent_cols:
                if c in row:
                    row[c] = _to_fraction(row[c])

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(formatted_rows, f, ensure_ascii=False, indent=2)
        logging.info(f"JSON exported: {file_path} ({len(formatted_rows)} rows)")
        return file_path
    except Exception as e:
        logging.error(f"Error exporting JSON for case {ticket_number}: {str(e)}")
        return None

def unique_model_numbers_and_serials(data_list):
    try:
        model_serials = {}

        if isinstance(data_list, dict):
            data_list = [data_list]

        for data in data_list:
            if isinstance(data, dict) and "Model number" in data and "Serial number" in data:
                model_value = data["Model number"]
                serial_value = data["Serial number"]

                if isinstance(model_value, list) and isinstance(serial_value, list):
                    for model, serial in zip(model_value, serial_value):
                        if model and model != 'Not available' and serial and serial != 'Not available':
                            if model not in model_serials:
                                model_serials[model] = serial
                elif model_value and model_value != 'Not available' and serial_value and serial_value != 'Not available':
                    if model_value not in model_serials:
                        model_serials[model_value] = serial_value

        return [[model, serial] for model, serial in model_serials.items()]
    except Exception as e:
        print(f"Error extracting model numbers and serials: {str(e)}")
        return []

def process_and_style_excel(file_path):
    # --- SELF-HEAL STEP A: ensure PM sheet name exists even if writer used default 'Sheet1' ---
    try:
        wb_tmp = openpyxl.load_workbook(file_path)
        MAIN_NAME = "Preventive Maintanance"
        if MAIN_NAME not in wb_tmp.sheetnames:
            # Rename the first sheet to PM
            first_sheet = wb_tmp[wb_tmp.sheetnames[0]]
            first_sheet.title = MAIN_NAME
            wb_tmp.save(file_path)
            logging.info("Renamed first sheet to 'Preventive Maintanance' during styling self-heal.")
    except Exception as e:
        logging.warning(f"Could not self-heal sheet name: {e}")
    # --- END SELF-HEAL STEP A ---

    try:
        df = pd.read_excel(file_path, engine='openpyxl', keep_default_na=False, na_values=[])
        logging.info("Excel file loaded successfully.")
    except Exception as e:
        logging.error(f"Failed to load Excel file: {e}")
        raise

    # --- DIAGNOSTIC: log suspicious IP cells without changing anything ---
    try:
        if "Interface ip address" in df.columns:
            import re as _re
            _ip_pat = _re.compile(r'^\s*(?:\d{1,3}\.){3}\d{1,3}\s*$')
            bad = []
            for i, v in enumerate(df["Interface ip address"].astype(str), start=2):  # +2 for Excel row numbers
                s = v.strip()
                if s not in {"Not available", "Unsupported IOS", "Require Manual Check"} and not _ip_pat.match(s):
                    bad.append((i, s))
            if bad:
                logging.warning(f"Suspicious IP values found (row, value): {bad[:10]}{'...' if len(bad)>10 else ''}")
    except Exception as _e:
        logging.debug(f"IP diagnostics skipped: {_e}")
    # --- END DIAGNOSTIC ---

    # --- safe accessor so index-based lookups don't crash if layout changes ---
    def _safe_iloc(row, idx, default=""):
        try:
            return row.iloc[idx] if 0 <= idx < len(row) else default
        except Exception:
            return default

    # --- Coerce percentage-like strings to numeric fractions for thresholding/formatting ---
    percentage_columns = ["CPU Utilization", "Memory Utilization (%)", "Used Flash (%)"]

    def _to_fraction(val):
        try:
            if isinstance(val, (int, float)):
                return float(val) if val <= 1 else val
            if isinstance(val, str):
                s = val.strip()
                m = re.match(r'^(\d+(?:\.\d+)?)\s*%$', s)
                if m:
                    return float(m.group(1)) / 100.0
        except Exception as e:
            logging.debug(f"Percentage coercion skipped for value '{val}': {e}")
        return val

    for col in percentage_columns:
        if col in df.columns:
            df[col] = df[col].apply(_to_fraction)

    # Recommendation functions (unchanged logic)
    def uptime(row):
        try:
            text = str(_safe_iloc(row, 5, ""))
            matches = re.findall(r'(\d+)\s+(year|week|day|hour|minute|second)s?', text)
            time_dict = {unit: int(value) for value, unit in matches}
            if time_dict.get('year', 0) > 1 or (
                time_dict.get('year', 0) == 1 and any(unit in time_dict for unit in ['week', 'day', 'hour', 'minute', 'second'])
            ):
                return "Consider to power cycle the device at the nearest maintenance window."
            total_days = (
                time_dict.get('week', 0) * 7 +
                time_dict.get('day', 0) +
                time_dict.get('hour', 0) / 24 +
                time_dict.get('minute', 0) / 1440 +
                time_dict.get('second', 0) / 86400
            )
            if total_days > 366:
                return "Consider to power cycle the device at the nearest maintenance window"
        except Exception as e:
            logging.warning(f"Error in uptime: {e}")
        return None

    def simple_check(row, index, trigger, message):
        try:
            value = str(_safe_iloc(row, index, "")).strip()
            if value == trigger:
                return message
        except Exception as e:
            logging.warning(f"Error in check at index {index}: {e}")
        return None

    def threshold_check(row, index, threshold, message):
        try:
            val = _safe_iloc(row, index, None)
            if isinstance(val, (int, float)) and val >= threshold:
                return message
        except Exception as e:
            logging.warning(f"Error in threshold check at index {index}: {e}")
        return None

    def psu_check(row):
        try:
            value = str(_safe_iloc(row, 25, "")).strip()
            if value != "OK":
                return "PSU functionalities are abnormal, try to reseat the PSU and verify the status."
        except Exception as e:
            logging.warning(f"Error in PSU check: {e}")
        return None

    def fan_check(row):
        try:
            value = str(_safe_iloc(row, 23, "")).strip()
            if value != "OK":
                return "Error noticed in fan functionality, kindly review."
        except Exception as e:
            logging.warning(f"Error in fan check: {e}")
        return None

    def temperature_check(row):
        try:
            value = str(_safe_iloc(row, 24, "")).strip()
            if value != "OK":
                return "Abnormalities noticed in device temperature, suggested to check the fan status and also room temperature if required."
        except Exception as e:
            logging.warning(f"Error in temperature check: {e}")
        return None

    def hardware_recommendations(row):
        try:
            today = datetime.today()
            one_year_later = today + timedelta(days=365)
            has_passed = is_approaching = False
            for i in range(27, 32):
                try:
                    date = parser.parse(str(_safe_iloc(row, i, "")), fuzzy=True)
                    if date < today:
                        has_passed = True
                    elif today <= date <= one_year_later:
                        is_approaching = True
                except (ValueError, TypeError):
                    continue
            if has_passed and is_approaching:
                return "One of the EOS milestones has already passed for the device model, please consider a hardware refresh."
            if has_passed:
                return "Device has already passed the last date of support from vendor, please consider hardware refresh."
            if is_approaching:
                return "Device is approaching the EOS soon, please consider a hardware refresh."
        except Exception as e:
            logging.warning(f"Error in hardware_recommendations: {e}")
        return None

    def duplex_check(row):
        return simple_check(row, 32, "YES", "Enable full duplex mode on all applicable interfaces to prevent performance issues.")

    def config_check(row):
        return simple_check(row, 34, "YES", "Unsaved configuration detected, recommended to save configurations to prevent loss during reboot.")

    def logs_check(row):
        return simple_check(row, 36, "YES", "Critical logs found in the device, please review.")

    def debug_check(row):
        return simple_check(row, 13, "YES", "The debug is enabled, please review the debug configurations and disable it as needed.")

    def memory_check(row):
        return threshold_check(row, 18, 0.8, "Memory utilization is found to be high, please review top processes consuming more memory.")

    def flash_check(row):
        return threshold_check(row, 22, 0.8, "Flash memory utilization is observed to be high, kindly review the top processes or files contributing to elevated flash usage.")

    def generate_comment(row):
        comments = []
        for func in [
            uptime, debug_check, memory_check, flash_check,
            psu_check, fan_check, temperature_check,
            hardware_recommendations, duplex_check,
            config_check, logs_check
        ]:
            try:
                result = func(row)
                if result:
                    comments.append(result)
            except Exception as e:
                logging.warning(f"Error executing function: {e}")
        if not comments:
            return "Device operating with good parameters."
        return "\n".join(comments)

    # --- PRESERVE existing Remark; only fill when blank ---
    try:
        remark_col = "Remark"
        if remark_col not in df.columns:
            df[remark_col] = ""
        blank_mask = df[remark_col].astype(str).str.strip().eq("")
        df.loc[blank_mask, remark_col] = df[blank_mask].apply(generate_comment, axis=1)
        df.to_excel(file_path, index=False, engine='openpyxl')
        logging.info("Excel file updated with recommendations (preserving existing Remark).")
    except Exception as e:
        logging.error(f"Failed to update Excel file: {e}")
        raise

    # Post-processing: styling, formatting, highlighting
    try:
        wb = openpyxl.load_workbook(file_path)
        red_fill = PatternFill(start_color="bc4f5e", end_color="bc4f5e", fill_type="solid")
        purple_fill = PatternFill(start_color="CBC3E3", end_color="CBC3E3", fill_type="solid")
        center_wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)
        percentage_columns = ["CPU Utilization", "Memory Utilization (%)", "Used Flash (%)"]
        date_columns = [
            "s/w release date", "End-of-Sale Date: HW", "Last Date of Support: HW",
            "End of Routine Failure Analysis Date:  HW", "End of Vulnerability/Security Support: HW",
            "End of SW Maintenance Releases Date: HW"
        ]

        for sheet in wb.worksheets:
            header = [cell.value for cell in sheet[1]]
            percent_indexes = [header.index(col) for col in percentage_columns if col in header]
            date_indexes = [header.index(col) for col in date_columns if col in header]

            for cell in sheet[1]:
                cell.fill = purple_fill
                cell.alignment = center_wrap_align
                cell.font = bold_font

            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    val = cell.value
                    mark_red = False
                    if isinstance(val, str):
                        s = val.strip()
                        # Expanded variants & common typos, plus unified Non-IOS_XE marker
                        red_markers = {
                            "Unavailable",
                            "Invalid inner data format",
                            "Invalid data format",
                            "Check manually",
                            "Require Manual Check",
                            "Yet to check",
                            "Command not found",
                            "Command not found.",      # trailing period variant
                            "Not available",
                            "NA",
                            "Not avialable",          # common typo seen in remarks
                            "Non-IOS_XE",
                            "Unsupported IOS",
                        }
                        if s in red_markers or s.startswith("Error:"):
                            mark_red = True
                    if mark_red:
                        cell.fill = red_fill
                    cell.alignment = center_wrap_align

                for idx in percent_indexes:
                    cell = row[idx]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'

                for idx in date_indexes:
                    cell = row[idx]
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'mmmm dd, yyyy'

            for col in sheet.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                sheet.column_dimensions[col[0].column_letter].width = max_length + 2

        wb.save(file_path)
        logging.info("Post-processing completed: styling, formatting, highlighting.")
    except Exception as e:
        logging.error(f"Post-processing failed: {e}")
        raise

    # --- SELF-HEAL STEP B: ensure Summary exists even if caller skipped add_summary_sheet ---
    try:
        add_summary_sheet(file_path)
        logging.info("Summary sheet ensured from process_and_style_excel (self-heal).")
    except Exception as e:
        logging.warning(f"Self-heal summary creation failed: {e}")
    # --- END SELF-HEAL STEP B ---

def add_summary_sheet(file_path):
    """
    Ensure the first sheet is named 'Preventive Maintanance' and create/refresh a 'Summary' sheet
    placed second (index 1). Adds rollups and counts unified 'Non-IOS_XE' rows based on 'Remark'.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        import re

        wb = openpyxl.load_workbook(file_path)

        MAIN_NAME = "Preventive Maintanance"  # exact spelling by request
        SUMMARY_NAME = "Summary"

        # 1) Ensure main sheet is first and named as required
        if MAIN_NAME in wb.sheetnames:
            main = wb[MAIN_NAME]
            wb._sheets.remove(main)
            wb._sheets.insert(0, main)
        else:
            first_sheet = wb[wb.sheetnames[0]]
            first_sheet.title = MAIN_NAME
            main = first_sheet

        # 2) Remove old Summary (if any), then create new at index 1
        if SUMMARY_NAME in wb.sheetnames:
            del wb[SUMMARY_NAME]
        ws = wb.create_sheet(SUMMARY_NAME, 1)

        # 3) Read headers & data rows from main
        headers = [c.value for c in main[1]]
        rows = [[cell for cell in r] for r in main.iter_rows(min_row=2, values_only=True)]

        # Helper: get column index by name
        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return -1

        idx_cpu   = col("CPU Utilization")
        idx_mem   = col("Memory Utilization (%)")
        idx_flash = col("Used Flash (%)")
        idx_fan   = col("Fan status")
        idx_temp  = col("Temperature status")
        idx_psu   = col("PowerSupply status")
        idx_fname = col("File name")
        idx_crit  = col("Critical logs")
        idx_rem   = col("Remark")

        def _as_frac(v):
            if isinstance(v, (int, float)):
                return float(v) if v <= 1 else None
            if isinstance(v, str):
                s = v.strip()
                m = re.match(r'^(\d+(?:\.\d+)?)\s*%$', s)
                if m:
                    return float(m.group(1)) / 100.0
            return None

        def avg(col_idx):
            vals = []
            if col_idx >= 0:
                for r in rows:
                    frac = _as_frac(r[col_idx])
                    if frac is not None:
                        vals.append(frac)
            return round(sum(vals) / len(vals), 4) if vals else None

        def count_status(col_idx):
            counts = {"OK": 0, "Not OK": 0, "Unsupported version": 0, "Not available": 0, "Other": 0}
            if col_idx >= 0:
                for r in rows:
                    v = r[col_idx]
                    if isinstance(v, list):
                        norm = "OK" if all(str(x).strip().upper() == "OK" for x in v) else "Not OK"
                    else:
                        norm = str(v).strip()
                    u = norm.upper()
                    if u == "OK":
                        counts["OK"] += 1
                    elif u in ("NOT OK", "NOT_OK", "NOTOK"):
                        counts["Not OK"] += 1
                    elif u in ("UNSUPPORTED VERSION", "UNSUPPORTED IOS"):  # map Unsupported IOS here
                        counts["Unsupported version"] += 1
                    elif u in ("NOT AVAILABLE", "NA"):
                        counts["Not available"] += 1
                    else:
                        counts["Other"] += 1
            return counts

        # Base rollups
        total_rows = len(rows)
        stacks = sum(1 for r in rows if idx_fname >= 0 and "_Stack_" in str(r[idx_fname]))
        singles = total_rows - stacks

        fan_c  = count_status(idx_fan)
        temp_c = count_status(idx_temp)
        psu_c  = count_status(idx_psu)
        crit_yes = sum(1 for r in rows if idx_crit >= 0 and str(r[idx_crit]).strip().upper() == "YES")

        # Unified Non-IOS_XE count (robust to spacing/case/old variants)
        def _norm(s):
            if s is None:
                return ""
            # normalize case; remove spaces, dashes, underscores, colons & parentheses
            return (
                str(s).upper()
                .replace(" ", "")
                .replace("-", "")
                .replace("_", "")
                .replace(":", "")
                .replace("(", "")
                .replace(")", "")
                .strip()
            )

        non_iosxe = 0
        if idx_rem >= 0:
            for r in rows:
                raw = r[idx_rem] if idx_rem < len(r) else ""
                norm = _norm(raw)
                # Count our new tag
                if "NONIOSXE" in norm:
                    non_iosxe += 1
                # Backward-compatibility with older labels if they still exist in the sheet
                elif norm.startswith("INFONONIOSXESKIPPED") or norm.startswith("INFONONIOSXE"):
                    non_iosxe += 1
                elif norm.startswith("ERRORCOULDNOTREADFILE") or norm.startswith("ERRORPROCESSINGFAILURE"):
                    non_iosxe += 1

        # Write Summary table
        summary = [
            ["Metric", "Value"],
            ["Total rows exported", total_rows],
            ["Single members", singles],
            ["Stack members", stacks],
            ["Avg CPU Utilization", avg(idx_cpu)],
            ["Avg Memory Utilization (%)", avg(idx_mem)],
            ["Avg Used Flash (%)", avg(idx_flash)],
            ["Critical logs (YES)", crit_yes],
            [],
            ["Fan — OK", fan_c["OK"]],
            ["Fan — Not OK", fan_c["Not OK"]],
            ["Fan — Unsupported version", fan_c["Unsupported version"]],
            ["Fan — Not available", fan_c["Not available"]],
            ["Fan — Other", fan_c["Other"]],
            [],
            ["Temperature — OK", temp_c["OK"]],
            ["Temperature — Not OK", temp_c["Not OK"]],
            ["Temperature — Unsupported version", temp_c["Unsupported version"]],
            ["Temperature — Not available", temp_c["Not available"]],
            ["Temperature — Other", temp_c["Other"]],
            [],
            ["PSU — OK", psu_c["OK"]],
            ["PSU — Not OK", psu_c["Not OK"]],
            ["PSU — Unsupported version", psu_c["Unsupported version"]],
            ["PSU — Not available", psu_c["Not available"]],
            ["PSU — Other", psu_c["Other"]],
            [],
            ["Non-IOS-XE files (skipped)", non_iosxe],
        ]

        for r_idx, row in enumerate(summary, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto width for first two columns
        for col_idx in range(1, 3):
            max_len = 0
            for r in range(1, len(summary) + 1):
                val = ws.cell(row=r, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        wb.save(file_path)
        logging.info("Main/summary sheet order and naming enforced (with Non-IOS_XE counter).")
    except Exception as e:
        logging.warning(f"Failed to create/arrange sheets: {e}")

def main():

    try:
        file_path = r""
        directory_path = r""
        ticket_number = "SVR3456789"

        # 1) Parse all eligible files in the directory
        data = Cisco_IOS_XE.process_directory(directory_path)

        # 2) Create/append Excel (also styles & adds Summary)
        excel_path = append_to_excel(ticket_number, data)  # returns the file path
        print(excel_path)  # keep your original print

        # 3) Optional JSON export (disabled for now)
        # json_path = export_json(ticket_number, data, file_path=None)  # defaults to <ticket>_network_analysis.json
        # if not json_path:
        #     logging.warning("JSON export failed (continuing).")
        # else:
        #     print(json_path)

    except Exception as e:
        print(f"Error in main: {str(e)}")

if __name__ == "__main__":
    main()