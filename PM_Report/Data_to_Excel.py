# Data_to_Excel.py
# Fully instrumented with structured logging via logging_setup.configure_logging

from __future__ import annotations

import os
import re
import json
import pandas as pd
from datetime import datetime, timedelta
from dateutil import parser
from tempfile import NamedTemporaryFile
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# central logging initializer (shared)
from logging_setup import configure_logging
logger = configure_logging(__file__)

# Try both package and flat import styles for Cisco_IOS_XE to avoid import headaches.
try:
    # when used as a package (e.g., from Switching or PM_Report)
    from . import Cisco_IOS_XE  # type: ignore
    logger.debug("Imported Cisco_IOS_XE via relative import (package mode).")
except Exception:
    try:
        import Cisco_IOS_XE  # type: ignore
        logger.debug("Imported Cisco_IOS_XE via flat import (script mode).")
    except Exception as e:
        logger.error(f"Unable to import Cisco_IOS_XE: {e}")

# ----- version banner so you can confirm the right module is loaded -----
PHASE2_VERSION = "Phase2-final-2025-08-17.hotfix1"
try:
    logger.info(f"[Data_to_Excel] Loaded module version: {PHASE2_VERSION}")
except Exception:
    pass
# -----------------------------------------------------------------------

MAIN_SHEET = "Preventive Maintanance"  # exact spelling as requested
SUMMARY_SHEET = "Summary"


# ==========================
# Small helpers
# ==========================

def _unwrap_value(val):
    while isinstance(val, list) and len(val) == 1:
        val = val[0]
    return val

def _safe_save_wb(wb: openpyxl.Workbook, path: str) -> None:
    """
    Atomic save to avoid partial/corrupt archives.
    """
    logger.info(f"_safe_save_wb: start path={path}")
    parent = os.path.dirname(path)
    if parent:
        try:
            os.makedirs(parent, exist_ok=True)
        except Exception as e:
            logger.warning(f"_safe_save_wb: could not ensure parent dir {parent}: {e}")
    with NamedTemporaryFile(delete=False, dir=parent or None, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
        logger.debug(f"_safe_save_wb: saved atomically to {path}")
    except Exception as e:
        logger.exception(f"_safe_save_wb: save/replace failed: {e}")
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        finally:
            raise


# ==========================
# Public API
# ==========================

def append_to_excel(ticket_number: str,
                    data_list,
                    file_path: Optional[str] = None,
                    column_order: Optional[List[str]] = None) -> Optional[str]:
    """
    Write (or append) device rows to Excel MAIN_SHEET, then style and add summary.
    """
    logger.info(f"append_to_excel: start ticket={ticket_number}, file_path={file_path}")
    if column_order is None:
        column_order = [
            'File name', 'Host name', 'Model number', 'Serial number', 'Interface ip address',
            'Uptime', 'Current s/w version', 'Current SW EOS', 'Suggested s/w ver', 's/w release date',
            'Latest S/W version', 'Production s/w is deffered or not?', 'Last Reboot Reason', 'Any Debug?',
            'CPU Utilization', 'Total memory', 'Used memory', 'Free memory', 'Memory Utilization (%)',
            'Total flash memory', 'Used flash memory', 'Free flash memory', 'Used Flash (%)',
            'Fan status', 'Temperature status', 'PowerSupply status', 'Available Free Ports',
            'End-of-Sale Date: HW', 'Last Date of Support: HW', 'End of Routine Failure Analysis Date:  HW',
            'End of Vulnerability/Security Support: HW', 'End of SW Maintenance Releases Date: HW',
            'Any Half Duplex', 'Interface/Module Remark', 'Config Status', 'Config Save Date',
            'Critical logs', 'Remark'
        ]
    if file_path is None:
        file_path = f"{ticket_number}_network_analysis.xlsx"

    # Normalize input into list of dict rows (wide dicts with list values)
    formatted_data: List[Dict[str, object]] = []
    if isinstance(data_list, dict):
        data_list = [data_list]

    logger.debug(f"append_to_excel: incoming items={len(data_list) if data_list else 0}")

    for idx, data in enumerate(data_list or []):
        if not isinstance(data, dict):
            logger.error(f"append_to_excel: skipping invalid data at index {idx}: {type(data)}")
            continue
        if not data:
            logger.debug(f"append_to_excel: skipping empty dict at index {idx}")
            continue

        # Determine row length by first list-like field
        data_length = 1
        for value in data.values():
            if isinstance(value, list) and len(value) > 0:
                data_length = len(value)
                break

        for i in range(data_length):
            row_data: Dict[str, object] = {}
            for key in column_order:
                if key in data:
                    value = data[key]
                    if isinstance(value, list):
                        use_val = value[i] if i < len(value) else 'Not available'
                        row_data[key] = _unwrap_value(use_val)
                    else:
                        row_data[key] = _unwrap_value(value)
                else:
                    row_data[key] = 'Not available'
            formatted_data.append(row_data)

    if not formatted_data:
        logger.warning("append_to_excel: no rows to write; returning None.")
        return None

    df = pd.DataFrame(formatted_data)
    try:
        df = df[column_order]
    except Exception as e:
        logger.warning(f"append_to_excel: column reorder failed (using as-is). Error: {e}")

    # Ensure parent folder exists
    try:
        parent = os.path.dirname(file_path)
        if parent and not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
            logger.debug(f"append_to_excel: created parent dir {parent}")
    except Exception as e:
        logger.warning(f"append_to_excel: could not ensure parent folder for Excel: {e}")

    # Write/append to MAIN_SHEET
    try:
        if os.path.exists(file_path):
            logger.info(f"append_to_excel: appending to existing file {file_path}")
            try:
                existing_df = pd.read_excel(file_path, sheet_name=MAIN_SHEET)
                logger.debug(f"append_to_excel: read existing MAIN_SHEET rows={len(existing_df)}")
            except Exception as e:
                logger.warning(f"append_to_excel: MAIN_SHEET missing/corrupt, reading first sheet. {e}")
                existing_df = pd.read_excel(file_path)
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                combined_df.to_excel(writer, index=False, sheet_name=MAIN_SHEET)
            logger.debug(f"append_to_excel: wrote combined rows={len(combined_df)}")
        else:
            logger.info(f"append_to_excel: creating new file {file_path}")
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, index=False, sheet_name=MAIN_SHEET)
            logger.debug(f"append_to_excel: wrote rows={len(df)} (new workbook)")
    except Exception as e:
        logger.exception(f"append_to_excel: error writing Excel: {e}")
        return None

    # Style + remarks + summary (best-effort)
    try:
        process_and_style_excel(file_path)
    except Exception as e:
        logger.warning(f"append_to_excel: styling step failed: {e}")

    try:
        add_summary_sheet(file_path)
    except Exception as e:
        logger.warning(f"append_to_excel: summary sheet step failed: {e}")

    logger.info(f"append_to_excel: done file={file_path}")
    return file_path


def export_json(ticket_number: str,
                data_list,
                file_path: Optional[str] = None,
                column_order: Optional[List[str]] = None,
                coerce_percentages: bool = True) -> Optional[str]:
    """
    Export rows to JSON. Optionally coerce percent columns to fractions (0..1).
    """
    logger.info(f"export_json: start ticket={ticket_number}, file_path={file_path}")
    if column_order is None:
        column_order = [
            'File name', 'Host name', 'Model number', 'Serial number', 'Interface ip address',
            'Uptime', 'Current s/w version', 'Current SW EOS', 'Suggested s/w ver', 's/w release date',
            'Latest S/W version', 'Production s/w is deffered or not?', 'Last Reboot Reason', 'Any Debug?',
            'CPU Utilization', 'Total memory', 'Used memory', 'Free memory', 'Memory Utilization (%)',
            'Total flash memory', 'Used flash memory', 'Free flash memory', 'Used Flash (%)',
            'Fan status', 'Temperature status', 'PowerSupply status', 'Available Free Ports',
            'End-of-Sale Date: HW', 'Last Date of Support: HW', 'End of Routine Failure Analysis Date:  HW',
            'End of Vulnerability/Security Support: HW', 'End of SW Maintenance Releases Date: HW',
            'Any Half Duplex', 'Interface/Module Remark', 'Config Status', 'Config Save Date',
            'Critical logs', 'Remark'
        ]

    if file_path is None:
        file_path = f"{ticket_number}_network_analysis.json"

    formatted_rows: List[Dict[str, object]] = []
    if isinstance(data_list, dict):
        data_list = [data_list]

    # Normalize into row objects
    for idx, data in enumerate(data_list or []):
        if not isinstance(data, dict) or not data:
            logger.warning(f"export_json: skipping invalid/empty item at index {idx}: {type(data)}")
            continue

        data_length = 1
        for v in data.values():
            if isinstance(v, list) and len(v) > 0:
                data_length = len(v)
                break

        for i in range(data_length):
            row_obj: Dict[str, object] = {}
            for key in column_order:
                if key in data:
                    value = data[key]
                    if isinstance(value, list):
                        val = _unwrap_value(value[i] if i < len(value) else 'Not available')
                    else:
                        val = _unwrap_value(value)
                else:
                    val = 'Not available'
                row_obj[key] = val
            formatted_rows.append(row_obj)

    # Optional coercion
    if coerce_percentages:
        percent_cols = ["CPU Utilization", "Memory Utilization (%)", "Used Flash (%)"]

        def _to_fraction(val):
            try:
                if isinstance(val, (int, float)):
                    return float(val) if val <= 1 else val
                if isinstance(val, str):
                    s = val.strip()
                    m = re.match(r'^(\d+(?:\.\d+)?)\s*%$', s)
                    if m:
                        return float(m.group(1)) / 100.0
            except Exception:
                pass
            return val

        for r in formatted_rows:
            for c in percent_cols:
                if c in r:
                    r[c] = _to_fraction(r[c])

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(formatted_rows, f, ensure_ascii=False, indent=2)
        logger.info(f"export_json: wrote rows={len(formatted_rows)} → {file_path}")
        return file_path
    except Exception as e:
        logger.exception(f"export_json: error writing JSON: {e}")
        return None


def unique_model_numbers_and_serials(data_list) -> List[List[str]]:
    logger.info("unique_model_numbers_and_serials: start")
    try:
        model_serials: Dict[str, str] = {}
        if isinstance(data_list, dict):
            data_list = [data_list]
        for idx, data in enumerate(data_list or []):
            if not isinstance(data, dict):
                continue
            if "Model number" not in data or "Serial number" not in data:
                continue
            mval = data["Model number"]
            sval = data["Serial number"]
            if isinstance(mval, list) and isinstance(sval, list):
                for model, serial in zip(mval, sval):
                    if model and model != 'Not available' and serial and serial != 'Not available':
                        model_serials.setdefault(str(model), str(serial))
            else:
                if mval and mval != 'Not available' and sval and sval != 'Not available':
                    model_serials.setdefault(str(mval), str(sval))
        out = [[model, serial] for model, serial in model_serials.items()]
        logger.debug(f"unique_model_numbers_and_serials: pairs={len(out)}")
        return out
    except Exception as e:
        logger.exception(f"unique_model_numbers_and_serials: exception {e}")
        return []


# ==========================
# Styling / Post-processing
# ==========================

def process_and_style_excel(file_path: str) -> None:
    """
    - Ensure main sheet exists/named correctly.
    - Update 'Remark' IN-PLACE (no sheet delete).
    - Style headers/cells and format percentage/date columns.
    - Save atomically.
    """
    logger.info(f"process_and_style_excel: start {file_path}")
    wb = openpyxl.load_workbook(file_path)

    # Ensure main sheet
    if MAIN_SHEET in wb.sheetnames:
        ws = wb[MAIN_SHEET]
    else:
        first = wb[wb.sheetnames[0]]
        first.title = MAIN_SHEET
        ws = first
        logger.debug("process_and_style_excel: renamed first sheet to MAIN_SHEET")

    headers = [c.value for c in ws[1]]
    name_to_idx = {name: idx for idx, name in enumerate(headers)}
    logger.debug(f"process_and_style_excel: header count={len(headers)}")

    def _safe_cell(r, c):
        try:
            return ws.cell(row=r, column=c+1).value
        except Exception:
            return ""

    def _as_percent_fraction(val):
        try:
            if isinstance(val, (int, float)):
                return float(val) if val <= 1 else None
            if isinstance(val, str):
                s = val.strip()
                m = re.match(r'^(\d+(?:\.\d+)?)\s*%$', s)
                if m:
                    return float(m.group(1)) / 100.0
        except Exception:
            return None
        return None

    def col(name: str) -> int:
        return name_to_idx.get(name, -1)

    idx_uptime = col("Uptime")

    def build_row_accessor(r):
        def _iloc(i):
            if i < 0 or i >= len(headers):
                return ""
            return _safe_cell(r, i)
        return _iloc

    # remark logic
    def uptime_comment(_iloc):
        try:
            text = str(_iloc(idx_uptime))
            matches = re.findall(r'(\d+)\s+(year|week|day|hour|minute|second)s?', text)
            time_dict = {unit: int(value) for value, unit in matches}
            if time_dict.get('year', 0) > 1 or (
                time_dict.get('year', 0) == 1 and any(unit in time_dict for unit in ['week', 'day', 'hour', 'minute', 'second'])
            ):
                return "Consider to power cycle the device at the nearest maintenance window."
            total_days = (
                time_dict.get('week', 0) * 7 +
                time_dict.get('day', 0) +
                time_dict.get('hour', 0) / 24 +
                time_dict.get('minute', 0) / 1440 +
                time_dict.get('second', 0) / 86400
            )
            if total_days > 366:
                return "Consider to power cycle the device at the nearest maintenance window"
        except Exception:
            pass
        return None

    def simple_check(_iloc, index_name, trigger, message):
        i = col(index_name)
        if i < 0:
            return None
        try:
            value = str(_iloc(i)).strip()
            if value == trigger:
                return message
        except Exception:
            pass
        return None

    def threshold_check(_iloc, index_name, threshold, message):
        i = col(index_name)
        if i < 0:
            return None
        try:
            v = _iloc(i)
            if isinstance(v, (int, float)):
                frac = v if v <= 1 else None
            else:
                frac = _as_percent_fraction(v)
            if frac is not None and frac >= threshold:
                return message
        except Exception:
            pass
        return None

    def psu_check(_iloc):
        return simple_check(_iloc, "PowerSupply status", "OK", None) or \
               ("PSU functionalities are abnormal, try to reseat the PSU and verify the status."
                if str(_iloc(col("PowerSupply status"))).strip() not in ("OK",) else None)

    def fan_check(_iloc):
        return simple_check(_iloc, "Fan status", "OK", None) or \
               ("Error noticed in fan functionality, kindly review."
                if str(_iloc(col("Fan status"))).strip() not in ("OK",) else None)

    def temperature_check(_iloc):
        return simple_check(_iloc, "Temperature status", "OK", None) or \
               ("Abnormalities noticed in device temperature, suggested to check the fan status and also room temperature if required."
                if str(_iloc(col("Temperature status"))).strip() not in ("OK",) else None)

    def hardware_recommendations(_iloc):
        try:
            today = datetime.today()
            one_year_later = today + timedelta(days=365)
            has_passed = is_approaching = False
            for name in ["End-of-Sale Date: HW", "Last Date of Support: HW",
                         "End of Routine Failure Analysis Date:  HW",
                         "End of Vulnerability/Security Support: HW",
                         "End of SW Maintenance Releases Date: HW"]:
                i = col(name)
                if i < 0:
                    continue
                val = _iloc(i)
                try:
                    date = parser.parse(str(val), fuzzy=True)
                    if date < today:
                        has_passed = True
                    elif today <= date <= one_year_later:
                        is_approaching = True
                except Exception:
                    continue
            if has_passed and is_approaching:
                return "One of the EOS milestones has already passed for the device model, please consider a hardware refresh."
            if has_passed:
                return "Device has already passed the last date of support from vendor, please consider hardware refresh."
            if is_approaching:
                return "Device is approaching the EOS soon, please consider a hardware refresh."
        except Exception:
            pass
        return None

    def duplex_check(_iloc):
        return simple_check(_iloc, "Any Half Duplex", "YES",
                            "Enable full duplex mode on all applicable interfaces to prevent performance issues.")

    def config_check(_iloc):
        return simple_check(_iloc, "Config Status", "YES",
                            "Unsaved configuration detected, recommended to save configurations to prevent loss during reboot.")

    def logs_check(_iloc):
        return simple_check(_iloc, "Critical logs", "YES",
                            "Critical logs found in the device, please review.")

    def debug_check(_iloc):
        return simple_check(_iloc, "Any Debug?", "YES",
                            "The debug is enabled, please review the debug configurations and disable it as needed.")

    def memory_check(_iloc):
        return threshold_check(_iloc, "Memory Utilization (%)", 0.8,
                               "Memory utilization is found to be high, please review top processes consuming more memory.")

    def flash_check(_iloc):
        return threshold_check(_iloc, "Used Flash (%)", 0.8,
                               "Flash memory utilization is observed to be high, kindly review the top processes or files contributing to elevated flash usage.")

    # Ensure Remark column exists
    remark_ci = name_to_idx.get("Remark", -1)
    if remark_ci < 0:
        remark_ci = len(headers)
        ws.cell(row=1, column=remark_ci+1, value="Remark")
        headers.append("Remark")
        name_to_idx["Remark"] = remark_ci
        logger.debug("process_and_style_excel: added missing Remark column")

    placeholders = {"", "Yet to check", "Not available", "NA", "Unsupported IOS"}
    placeholder_fold = {p.casefold() for p in placeholders}

    # Fill or keep Remark
    rows_filled = 0
    for r in range(2, ws.max_row + 1):
        _iloc = build_row_accessor(r)
        cur_val = ws.cell(row=r, column=remark_ci+1).value
        s = (str(cur_val).strip() if cur_val is not None else "")
        if s == "" or s.casefold() in placeholder_fold or s.startswith("Error:"):
            comments = []
            for fn in (uptime_comment, debug_check, memory_check, flash_check,
                       psu_check, fan_check, temperature_check,
                       hardware_recommendations, duplex_check,
                       config_check, logs_check):
                try:
                    msg = fn(_iloc)
                    if msg:
                        comments.append(msg)
                except Exception:
                    continue
            ws.cell(row=r, column=remark_ci+1, value=("\n".join(comments) if comments else "Device operating with good parameters."))
            rows_filled += 1
    logger.debug(f"process_and_style_excel: remark rows updated={rows_filled}")

    # Styling and formats
    red_fill = PatternFill(start_color="bc4f5e", end_color="bc4f5e", fill_type="solid")
    purple_fill = PatternFill(start_color="CBC3E3", end_color="CBC3E3", fill_type="solid")
    center_wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)

    percentage_columns = ["CPU Utilization", "Memory Utilization (%)", "Used Flash (%)"]
    date_columns = [
        "s/w release date", "End-of-Sale Date: HW", "Last Date of Support: HW",
        "End of Routine Failure Analysis Date:  HW", "End of Vulnerability/Security Support: HW",
        "End of SW Maintenance Releases Date: HW"
    ]

    # Style all sheets
    for sheet in wb.worksheets:
        header_cells = [cell.value for cell in sheet[1]]
        pct_idx = [header_cells.index(c) for c in percentage_columns if c in header_cells]
        date_idx = [header_cells.index(c) for c in date_columns if c in header_cells]

        # header style
        for cell in sheet[1]:
            cell.fill = purple_fill
            cell.alignment = center_wrap_align
            cell.font = bold_font

        # cell styling + conditional red fill
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                val = cell.value
                mark_red = False
                if isinstance(val, str):
                    s = val.strip()
                    red_markers = {
                        "Unavailable",
                        "Invalid inner data format",
                        "Invalid data format",
                        "Check manually",
                        "Require Manual Check",
                        "Yet to check",
                        "Command not found",
                        "Command not found.",
                        "NA",
                        "Non-IOS_XE",
                        "Unsupported IOS",
                        "Unsupported version",
                    }
                    if s in red_markers or s.startswith("Error:"):
                        mark_red = True
                if mark_red:
                    cell.fill = red_fill
                cell.alignment = center_wrap_align

            # percentage formats
            for idx in pct_idx:
                c = row[idx]
                if isinstance(c.value, (int, float)):
                    c.number_format = '0.00%'

            # date formats
            for idx in date_idx:
                c = row[idx]
                if isinstance(c.value, datetime):
                    c.number_format = 'mmmm dd, yyyy'

        # Auto width
        for col_cells in sheet.columns:
            try:
                max_length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
                sheet.column_dimensions[col_cells[0].column_letter].width = max_length + 2
            except Exception:
                # keep going if a column has no letter (merged?), very rare
                continue

    _safe_save_wb(wb, file_path)
    logger.info("process_and_style_excel: completed & saved.")


def add_summary_sheet(file_path: str) -> None:
    """
    Ensure MAIN_SHEET is first and create/refresh a 'Summary' sheet.
    """
    logger.info(f"add_summary_sheet: start {file_path}")
    wb = openpyxl.load_workbook(file_path)

    # Ensure main sheet first and correctly named
    if MAIN_SHEET in wb.sheetnames:
        main = wb[MAIN_SHEET]
        wb._sheets.remove(main)
        wb._sheets.insert(0, main)
    else:
        first_sheet = wb[wb.sheetnames[0]]
        first_sheet.title = MAIN_SHEET
        main = first_sheet
        logger.debug("add_summary_sheet: renamed first sheet to MAIN_SHEET")

    # Recreate Summary
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET, 1)

    headers = [c.value for c in main[1]]
    rows = [[cell for cell in r] for r in main.iter_rows(min_row=2, values_only=True)]
    logger.debug(f"add_summary_sheet: main rows={len(rows)}, headers={len(headers)}")

    def col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    idx_cpu   = col("CPU Utilization")
    idx_mem   = col("Memory Utilization (%)")
    idx_flash = col("Used Flash (%)")
    idx_fan   = col("Fan status")
    idx_temp  = col("Temperature status")
    idx_psu   = col("PowerSupply status")
    idx_fname = col("File name")
    idx_crit  = col("Critical logs")
    idx_rem   = col("Remark")

    def _as_frac(v):
        if isinstance(v, (int, float)):
            return float(v) if v <= 1 else None
        if isinstance(v, str):
            s = v.strip()
            m = re.match(r'^(\d+(?:\.\d+)?)\s*%$', s)
            if m:
                return float(m.group(1)) / 100.0
        return None

    def avg(col_idx):
        vals = []
        if col_idx >= 0:
            for r in rows:
                frac = _as_frac(r[col_idx])
                if frac is not None:
                    vals.append(frac)
        return round(sum(vals) / len(vals), 4) if vals else None

    def count_status(col_idx):
        counts = {"OK": 0, "Not OK": 0, "Unsupported version": 0, "Not available": 0, "Other": 0}
        if col_idx >= 0:
            for r in rows:
                v = r[col_idx]
                # normalize list-of-status (e.g., fans [..]) into a single "OK" if all ok
                norm = "OK" if isinstance(v, list) and all(str(x).strip().upper() == "OK" for x in v) else str(v).strip()
                u = norm.upper()
                if u == "OK":
                    counts["OK"] += 1
                elif u in ("NOT OK", "NOT_OK", "NOTOK"):
                    counts["Not OK"] += 1
                elif u in ("UNSUPPORTED VERSION", "UNSUPPORTED IOS"):
                    counts["Unsupported version"] += 1
                elif u in ("NOT AVAILABLE", "NA"):
                    counts["Not available"] += 1
                else:
                    counts["Other"] += 1
        return counts

    total_rows = len(rows)
    stacks = sum(1 for r in rows if idx_fname >= 0 and "_Stack_" in str(r[idx_fname]))
    singles = total_rows - stacks
    logger.debug(f"add_summary_sheet: singles={singles}, stacks={stacks}")

    fan_c  = count_status(idx_fan)
    temp_c = count_status(idx_temp)
    psu_c  = count_status(idx_psu)
    crit_yes = sum(1 for r in rows if idx_crit >= 0 and str(r[idx_crit]).strip().upper() == "YES")

    def _norm(s):
        if s is None:
            return ""
        return (
            str(s).upper()
            .replace(" ", "")
            .replace("-", "")
            .replace("_", "")
            .replace(":", "")
            .replace("(", "")
            .replace(")", "")
            .strip()
        )

    non_iosxe = 0
    if idx_rem >= 0:
        for r in rows:
            norm = _norm(r[idx_rem] if idx_rem < len(r) else "")
            if "NONIOSXE" in norm:
                non_iosxe += 1
            elif norm.startswith("INFONONIOSXESKIPPED") or norm.startswith("INFONONIOSXE"):
                non_iosxe += 1
            elif norm.startswith("ERRORCOULDNOTREADFILE") or norm.startswith("ERRORPROCESSINGFAILURE"):
                non_iosxe += 1

    summary = [
        ["Metric", "Value"],
        ["Total rows exported", total_rows],
        ["Single members", singles],
        ["Stack members", stacks],
        ["Avg CPU Utilization", avg(idx_cpu)],
        ["Avg Memory Utilization (%)", avg(idx_mem)],
        ["Avg Used Flash (%)", avg(idx_flash)],
        ["Critical logs (YES)", crit_yes],
        [],
        ["Fan — OK", fan_c["OK"]],
        ["Fan — Not OK", fan_c["Not OK"]],
        ["Fan — Unsupported version", fan_c["Unsupported version"]],
        ["Fan — Not available", fan_c["Not available"]],
        ["Fan — Other", fan_c["Other"]],
        [],
        ["Temperature — OK", temp_c["OK"]],
        ["Temperature — Not OK", temp_c["Not OK"]],
        ["Temperature — Unsupported version", temp_c["Unsupported version"]],
        ["Temperature — Not available", temp_c["Not available"]],
        ["Temperature — Other", temp_c["Other"]],
        [],
        ["PSU — OK", psu_c["OK"]],
        ["PSU — Not OK", psu_c["Not OK"]],
        ["PSU — Unsupported version", psu_c["Unsupported version"]],
        ["PSU — Not available", psu_c["Not available"]],
        ["PSU — Other", psu_c["Other"]],
        [],
        ["Non-IOS-XE files (skipped)", non_iosxe],
    ]

    for r_idx, row in enumerate(summary, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # auto width first two columns
    for col_idx in range(1, 3):
        try:
            max_len = 0
            for r in range(1, len(summary) + 1):
                val = ws.cell(row=r, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        except Exception:
            continue

    _safe_save_wb(wb, file_path)
    logger.info("add_summary_sheet: completed & saved.")


# ==========================
# Optional local main
# ==========================

def main():
    """
    Local test entry:
    - parse directory through Cisco_IOS_XE.process_directory
    - append to Excel
    """
    logger.info("Data_to_Excel.main: start")
    try:
        directory_path = r""  # fill if you want a quick local test
        ticket_number = "SVR3456789"
        if not directory_path:
            logger.warning("Data_to_Excel.main: directory_path is empty → skipping.")
            return

        data = Cisco_IOS_XE.process_directory(directory_path)
        if isinstance(data, int):
            logger.error(f"Data_to_Excel.main: process_directory returned error code {data}")
            return

        excel_path = append_to_excel(ticket_number, data)  # also styles & adds Summary
        logger.info(f"Data_to_Excel.main: wrote Excel to {excel_path}")
    except Exception as e:
        logger.exception(f"Data_to_Excel.main: exception {e}")

if __name__ == "__main__":
    main()
